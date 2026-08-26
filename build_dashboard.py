"""
RCRC Executive Dashboard Auto-Builder
--------------------------------------
Reads Stakeholder Engagement Plan (4).xlsx and injects fresh data into
window.DASH_DATA inside index.html, WITHOUT touching any HTML, CSS, or
JavaScript function outside that single data object.

This script is designed to run inside GitHub Actions on every push,
so any Excel update automatically produces an updated index.html that
Netlify then deploys.
"""

import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

EXCEL_FILE = "Stakeholder Engagement Plan (4).xlsx"
HTML_FILE = "index.html"


def fmt(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def table_from_named_table(wb, sheet_name, table_name):
    """
    Read an Excel named table and return its rows as dictionaries.

    For tblPreviousTracker only:
    rows with a blank S.No are merged into the previous meeting record.
    Other tables remain unchanged.
    """

    ws = wb[sheet_name]
    tbl = ws.tables.get(table_name)

    if tbl is None:
        raise ValueError(
            f"Table '{table_name}' not found in sheet '{sheet_name}'"
        )

    cells = ws[tbl.ref]
    rows = [[cell.value for cell in row] for row in cells]

    if not rows:
        return []

    headers = [fmt(value) for value in rows[0]]

    # Make duplicate column headers unique
    seen = {}
    clean_headers = []

    for header in headers:
        if header in seen:
            seen[header] += 1
            clean_headers.append(f"{header}{seen[header]}")
        else:
            seen[header] = 1
            clean_headers.append(header)

    raw_records = []

    for row in rows[1:]:
        if all(
            value is None or str(value).strip() == ""
            for value in row
        ):
            continue

        record = {}

        for header, value in zip(clean_headers, row):
            if not header:
                continue

            record[header] = fmt(value)

        raw_records.append(record)

    # Do not merge rows in any other Excel table
    if table_name != "tblPreviousTracker":
        return raw_records

    merged_records = []
    current_record = None

    merge_fields = [
        "Key Point Discussed",
        "Action Items",
        "Minutes of meeting",
        "Minutes of meeting2",
        "Minutes of meeting3",
        "Minutes of meeting4",
        "Minutes of meeting5",
        "Minutes of meeting6",
        "Minutes of meeting7",
        "Minutes of meeting8",
        "follow-up",
        "their response note",
        "Our Actions",
    ]

    for record in raw_records:
        serial_number = (     str(record.get("S.No", "")).strip()     or str(record.get("S.no", "")).strip()     or str(record.get("S No", "")).strip()     or str(record.get("No", "")).strip() )

        # A non-empty S.No means a new meeting
        if serial_number:
            current_record = dict(record)
            merged_records.append(current_record)
            continue

        # Ignore orphan continuation rows
        if current_record is None:
            continue

        # Merge continuation content into the previous meeting
        for field in merge_fields:
            new_value = str(record.get(field, "")).strip()

            if not new_value:
                continue

            existing_value = str(
                current_record.get(field, "")
            ).strip()

            if existing_value:
                current_record[field] = (
                    existing_value + "\n" + new_value
                )
            else:
                current_record[field] = new_value

    return merged_records

def find_table_sheet(wb, table_name):
    for ws in wb.worksheets:
        if table_name in ws.tables:
            return ws.title
    raise ValueError(f"Table '{table_name}' not found in any sheet")


def main():
    wb = load_workbook(EXCEL_FILE, data_only=True)

    tables = {
        "schedulePriority": "tblSchedulingPriority",
        "risks": "tblRiskRegister",
        "previous": "tblPreviousTracker",
        "followUpStatus": "tblFollowUpStatus",
        "questions": "tblQuestionBank",
        "stakeholders": "tblEngagementStrategy",
    }

    data = {}
    for key, table_name in tables.items():
        sheet = find_table_sheet(wb, table_name)
        data[key] = table_from_named_table(wb, sheet, table_name)

if key == "previous":
    print(f"[DEBUG] previous records = {len(data[key])}")
        print(f"[ok] {key}: {len(data[key])} rows from '{table_name}' ({sheet})")

    if len(data["schedulePriority"]) == 0:
        print("[error] schedulePriority table is empty, aborting to protect dashboard")
        sys.exit(1)
    if len(data["risks"]) == 0:
        print("[error] risks table is empty, aborting to protect dashboard")
        sys.exit(1)

    html_path = Path(HTML_FILE)
    html = html_path.read_text(encoding="utf-8")

    pattern = re.compile(r"window\.DASH_DATA\s*=\s*(\{.*?\});", re.S)
    match = pattern.search(html)
    if not match:
        print("[error] Could not locate window.DASH_DATA block in index.html")
        sys.exit(1)

    try:
        existing = json.loads(match.group(1))
    except Exception as e:
        print(f"[error] Existing DASH_DATA is not valid JSON, aborting: {e}")
        sys.exit(1)

    existing.update(data)

    new_json = json.dumps(existing, ensure_ascii=False, separators=(",", ":"))
    new_block = "window.DASH_DATA = " + new_json + ";"

    new_html = html[: match.start()] + new_block + html[match.end():]

    if len(new_html) < len(html) * 0.5:
        print("[error] Resulting HTML looks truncated, aborting to protect dashboard")
        sys.exit(1)

    html_path.write_text(new_html, encoding="utf-8")
    print(f"[done] index.html updated successfully ({len(new_html)} bytes)")


if __name__ == "__main__":
    main()
